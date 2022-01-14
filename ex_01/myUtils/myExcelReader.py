from openpyxl import load_workbook

baseDIR = ("E:\learn\problem_01")

def readExcel(fileName, sheetName, sheet_X, Sheet_Y):
    book = load_workbook(baseDIR + "\\" + fileName)
    sheet = book[sheetName]
    return tuple(sheet[sheet_X:Sheet_Y])

def getCustomResult(fileName, sheetName,sheet_X,Sheet_Y):
    book = load_workbook(baseDIR + "\\" + fileName)
    sheet = book[sheetName]
    customData = []
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data for one census tract.
        title = sheet['A' + str(row)].value
        value = sheet['C' + str(row)].value
        customData.append((title,value))
    return customData